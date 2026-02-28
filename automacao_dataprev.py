#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automação - Validação Regulatória de Contratos (DATAPREV)
=========================================================
Descrição:
  Lê o arquivo .csv.gz disponibilizado quinzenalmente pelo órgão gestor,
  consulta o sistema interno via web scraping e gera um relatório Excel
  com a situação regulatória de cada contrato.

Fluxo:
  1. Lê o arquivo .csv.gz e extrai os números de contrato
  2. Para cada contrato, acessa o sistema interno e busca:
     - Fase da Esteira (etapa atual)
     - Data de Averbação (histórico da etapa Averbação)
  3. Gera relatório Excel com as colunas:
     Sequencial | Número_Contrato | Fase da Esteira | Averbação

Autora: Daiane Vinharski
"""

import gzip
import os
import sys
import time
import pandas as pd

# Força UTF-8 no terminal do Windows
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, ElementClickInterceptedException
)

# ============================================================
# CONFIGURAÇÕES — preencha com seus dados antes de executar
# ============================================================
# Caminho para o arquivo .csv.gz disponibilizado pelo órgão
ARQUIVO_GZ    = r"C:\caminho\para\arquivo.csv.gz"

# Caminho de saída do relatório Excel
ARQUIVO_SAIDA = r"C:\caminho\para\Relatorio_DATAPREV.xlsx"

# Caminho para o driver do navegador (Edge)
DRIVER_PATH   = r"C:\caminho\para\msedgedriver.exe"

# URL do sistema interno
URL_SISTEMA   = "https://seu-sistema-interno.com/propostas"

# Credenciais de acesso — recomendado usar variáveis de ambiente
# Exemplo: LOGIN_USER = os.environ.get("SISTEMA_USER")
LOGIN_USER    = os.environ.get("SISTEMA_USER", "seu_usuario")
LOGIN_PASS    = os.environ.get("SISTEMA_PASS", "sua_senha")

ESPERA_PADRAO           = 15   # segundos de timeout para WebDriverWait
PAUSA_ENTRE_CONTRATOS   = 1.0  # segundos entre cada consulta


# ============================================================
# PARTE 1 — Extração dos contratos do arquivo regulatório
# ============================================================
def extrair_contratos(caminho_arquivo: str) -> list[str]:
    """Retorna lista de números de contrato únicos extraídos do .csv.gz"""
    print("\n" + "="*60)
    print("ETAPA 1 — Lendo arquivo regulatório")
    print("="*60)

    if not os.path.exists(caminho_arquivo):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")

    contratos = []

    # Tentativa 1: Pandas (mais robusto)
    try:
        df = pd.read_csv(caminho_arquivo, sep=';', compression='gzip',
                         encoding='latin1', header=None, on_bad_lines='skip')
        contratos = df.iloc[:, -1].dropna().astype(str).str.strip().tolist()
        print(f"✓ Pandas: {len(contratos)} registros lidos")
    except Exception as e:
        print(f"✗ Pandas falhou: {e}")

    # Tentativa 2: gzip direto
    if not contratos:
        try:
            with gzip.open(caminho_arquivo, 'rb') as f:
                texto = f.read().decode('latin1')
            linhas = texto.split('\n')
            for linha in linhas:
                linha = linha.strip()
                if linha and ';' in linha:
                    partes = linha.split(';')
                    val = partes[-1].strip()
                    if val:
                        contratos.append(val)
                elif linha:
                    contratos.append(linha)
            print(f"✓ gzip direto: {len(contratos)} registros lidos")
        except Exception as e:
            print(f"✗ gzip direto falhou: {e}")

    if not contratos:
        raise RuntimeError("Não foi possível ler o arquivo. Verifique o formato.")

    # Limpeza: manter apenas valores numéricos
    contratos_validos = []
    ignorados = 0
    for c in contratos:
        valor = c.strip()
        if valor.isdigit():
            contratos_validos.append(valor)
        else:
            ignorados += 1

    # Remover duplicatas mantendo a ordem
    contratos = list(dict.fromkeys(contratos_validos))
    print(f"✓ Total de contratos únicos: {len(contratos)}")
    if ignorados:
        print(f"  ℹ {ignorados} linha(s) ignorada(s) (cabeçalhos ou valores não numéricos)")
    return contratos


# ============================================================
# PARTE 2 — Automação Selenium
# ============================================================
def iniciar_navegador() -> webdriver.Edge:
    """Abre o Edge e realiza o login no sistema interno"""
    print("\n" + "="*60)
    print("ETAPA 2 — Iniciando navegador e realizando login")
    print("="*60)

    options = webdriver.EdgeOptions()
    # options.add_argument("--headless")  # descomente para rodar sem janela
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")

    service = Service(executable_path=DRIVER_PATH)
    driver = webdriver.Edge(service=service, options=options)
    wait = WebDriverWait(driver, ESPERA_PADRAO)

    driver.get(URL_SISTEMA)

    # Preencher login — ajuste os seletores conforme o seu sistema
    wait.until(EC.presence_of_element_located((By.ID, 'login_username'))).send_keys(LOGIN_USER)
    time.sleep(1)
    driver.find_element(By.ID, 'login_password').send_keys(LOGIN_PASS)
    time.sleep(1)

    # Clicar no botão de entrar
    wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[.//span[contains(text(),'Entrar') or contains(text(),'Login')]] | //form//button[@type='submit']")
    )).click()

    # Aguardar carregar — ajuste o seletor para um elemento da tela pós-login
    wait.until(EC.presence_of_element_located(
        (By.XPATH, "//input[@placeholder='Pesquisa' or contains(@placeholder,'pesquisa')]")
    ))
    print("✓ Login realizado com sucesso")

    return driver


def buscar_proposta(driver: webdriver.Edge, numero_contrato: str) -> tuple[str, str]:
    """
    Busca um contrato no sistema e retorna (fase_esteira, data_averbacao).
    Retorna ('NÃO ENCONTRADO', '') se não localizar a proposta.
    """
    wait = WebDriverWait(driver, ESPERA_PADRAO)

    fase_esteira   = "NÃO ENCONTRADO"
    data_averbacao = ""

    try:
        # ── Pré-passo: fechar qualquer modal ainda aberto ────────────────
        try:
            modais = driver.find_elements(
                By.XPATH, "//button[contains(@class,'ant-modal-close')]"
            )
            for m in modais:
                try:
                    driver.execute_script("arguments[0].click();", m)
                except Exception:
                    pass
            if modais:
                time.sleep(0.4)
        except Exception:
            pass
        try:
            driver.switch_to.active_element.send_keys(Keys.ESCAPE)
        except Exception:
            pass
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.5)

        # ── Passo 1: Limpar campo de pesquisa e digitar o número ─────────
        campo_pesquisa = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//input[@placeholder='Pesquisa' or contains(@placeholder,'esquisa')]")
            )
        )
        # Limpeza via JavaScript (necessário para frameworks React/Angular)
        driver.execute_script("""
            var el = arguments[0];
            el.focus();
            var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(el, '');
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        """, campo_pesquisa)
        time.sleep(0.3)

        valor_atual = campo_pesquisa.get_attribute("value")
        if valor_atual:
            driver.execute_script("arguments[0].click();", campo_pesquisa)
            campo_pesquisa.send_keys(Keys.CONTROL + "a")
            time.sleep(0.1)
            campo_pesquisa.send_keys(Keys.DELETE)
            time.sleep(0.2)

        campo_pesquisa.send_keys(numero_contrato)
        time.sleep(0.3)
        campo_pesquisa.send_keys(Keys.RETURN)
        time.sleep(0.5)

        # ── Passo 2: Clicar em Filtrar ────────────────────────────────────
        btn_filtrar = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Filtrar')]"))
        )
        btn_filtrar.click()
        time.sleep(2)

        # ── Passo 3: Verificar resultado ──────────────────────────────────
        XPATHS_SEM_RESULTADO = (
            "//*[contains(text(),'Nenhum') or contains(text(),'nenhum')]"
            "[contains(text(),'registro') or contains(text(),'resultado') or contains(text(),'encontrado')]"
        )
        if driver.find_elements(By.XPATH, XPATHS_SEM_RESULTADO):
            print(f"  ℹ Contrato {numero_contrato}: nenhum resultado")
            return fase_esteira, data_averbacao

        try:
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//tbody/tr[td] | //div[contains(@class,'list')]//div[contains(@class,'item')]")
                )
            )
        except TimeoutException:
            print(f"  ⚠ Contrato {numero_contrato}: nenhum resultado retornado")
            return fase_esteira, data_averbacao

        # ── Passo 4: Coletar Fase da Esteira ─────────────────────────────
        FASES_VALIDAS = {
            'formalização digital', 'protocolar documentação', 'aprovação corban',
            'in100', 'crédito', 'formalização interna', 'averbação',
            'lançamento do título', 'pago', 'emissão de cartão',
            'integração documental', 'integrado', 'cancelado'
        }
        fase_esteira = "NÃO IDENTIFICADO"
        try:
            td2 = WebDriverWait(driver, 8).until(
                EC.presence_of_element_located((By.XPATH, "(//tbody/tr)[1]/td[2]"))
            )
            linhas = [l.strip() for l in td2.text.split('\n') if l.strip()]
            for linha in linhas:
                if linha.lower() in FASES_VALIDAS:
                    fase_esteira = linha
                    break
            if fase_esteira == "NÃO IDENTIFICADO" and linhas:
                status_ignorar = {'em andamento', 'cancelado', 'aprovado', 'reprovado',
                                  'pendente', 'aguardando', 'em análise'}
                for linha in linhas:
                    if linha.lower() not in status_ignorar and len(linha) > 3:
                        fase_esteira = linha
                        break
        except Exception as e:
            print(f"  ✗ Erro ao ler fase: {e}")

        print(f"  • Fase: {fase_esteira}")

        # ── Passo 5: Expandir a linha ─────────────────────────────────────
        seta_clicada = False
        tentativas_seta = [
            "(//tbody/tr)[1]/td[last()]//button",
            "(//tbody/tr)[1]/td[last()]",
            "(//tbody/tr)[1]//button[last()]",
            "(//tbody/tr)[1]//*[contains(@class,'chevron') or contains(@class,'expand') or contains(@class,'arrow')]/..",
            "(//tbody/tr)[1]//*[name()='svg']/..",
        ]
        for xpath_seta in tentativas_seta:
            try:
                seta = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, xpath_seta))
                )
                driver.execute_script("arguments[0].scrollIntoView(true);", seta)
                time.sleep(0.2)
                seta.click()
                time.sleep(1)
                seta_clicada = True
                print(f"  ✓ Linha expandida")
                break
            except Exception:
                continue

        if not seta_clicada:
            print(f"  ⚠ Botão de expandir não encontrado para {numero_contrato}")
            return fase_esteira, data_averbacao

        # ── Passo 6: Localizar e clicar no ícone de Averbação ────────────
        icone_clicado = False
        tentativas_icone = [
            "//*[normalize-space(text())='Averbação']/following::button[1]",
            "//*[normalize-space(text())='Averbação']/ancestor::div[1]//button",
            "//*[normalize-space(text())='Averbação']/ancestor::td[1]//button",
            "//*[contains(@class,'step') or contains(@class,'stage')][.//*[contains(text(),'Averbação')]]//button",
            "//*[contains(text(),'Averbação')]/../button",
            "//*[contains(text(),'Averbação')]/following-sibling::button",
        ]
        for xpath_icone in tentativas_icone:
            try:
                icone = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, xpath_icone))
                )
                driver.execute_script("arguments[0].scrollIntoView(true);", icone)
                time.sleep(0.2)
                icone.click()
                time.sleep(1)
                icone_clicado = True
                print(f"  ✓ Ícone Averbação clicado")
                break
            except Exception:
                continue

        if not icone_clicado:
            print(f"  ⚠ Ícone de Averbação não encontrado para {numero_contrato}")
            return fase_esteira, data_averbacao

        # ── Passo 7: Ler a data no histórico ─────────────────────────────
        try:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(),'Histórico')]")
            ))
            try:
                aba_historico = driver.find_element(
                    By.XPATH, "//button[contains(text(),'Histórico')] | //div[contains(@class,'tab')][contains(text(),'Histórico')]"
                )
                aba_historico.click()
                time.sleep(0.5)
            except NoSuchElementException:
                pass

            linha_aprovacao = wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "//*[contains(text(),'aprovada ao realizar averba') or contains(text(),'averbação na Dataprev')]")
                )
            )
            try:
                tr = linha_aprovacao.find_element(By.XPATH, "./ancestor::tr[1]")
                celulas = tr.find_elements(By.TAG_NAME, "td")
                if celulas:
                    data_averbacao = celulas[0].text.strip().split(' ')[0]
            except Exception:
                try:
                    data_averbacao = linha_aprovacao.find_element(
                        By.XPATH, "./preceding-sibling::td[1] | ../preceding-sibling::tr[1]/td[1]"
                    ).text.strip().split(' ')[0]
                except Exception:
                    pass

        except TimeoutException:
            print(f"  ⚠ Histórico de averbação não encontrado para {numero_contrato}")

        print(f"  • Averbação: {data_averbacao}")

        # ── Passo 8: Fechar o modal ───────────────────────────────────────
        fechou = False
        for xpath_fechar in [
            "//button[contains(@class,'ant-modal-close')]",
            "//button[contains(normalize-space(.),'Fechar')]",
            "//button[contains(@aria-label,'fechar') or contains(@aria-label,'close')]",
            "//button[contains(@class,'close')]",
        ]:
            try:
                btn = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, xpath_fechar))
                )
                driver.execute_script("arguments[0].click();", btn)
                fechou = True
                time.sleep(0.5)
                break
            except Exception:
                continue
        if not fechou:
            try:
                driver.switch_to.active_element.send_keys(Keys.ESCAPE)
            except Exception:
                pass
            time.sleep(0.5)

    except Exception as e:
        print(f"  ✗ Erro ao consultar {numero_contrato}: {e}")

    return fase_esteira, data_averbacao


# ============================================================
# PARTE 3 — Geração do relatório Excel
# ============================================================

CORES_FASE = {
    'integrado':               ('00B050', 'FFFFFF'),
    'pago':                    ('FFD700', '000000'),
    'emissão de cartão':       ('4472C4', 'FFFFFF'),
    'integração documental':   ('C00000', 'FFFFFF'),
    'averbação':               ('FF6600', 'FFFFFF'),
    'formalização digital':    ('7030A0', 'FFFFFF'),
    'formalização interna':    ('C9549A', 'FFFFFF'),
    'aprovação corban':        ('008B8B', 'FFFFFF'),
    'in100':                   ('00B0F0', 'FFFFFF'),
    'crédito':                 ('ED7D31', 'FFFFFF'),
    'protocolar documentação': ('5B5EA6', 'FFFFFF'),
    'lançamento do título':    ('843C0C', 'FFFFFF'),
    'cancelado':               ('595959', 'FFFFFF'),
    'não encontrado':          ('D9D9D9', '000000'),
    'não identificado':        ('BFBFBF', '000000'),
}


def salvar_excel(dados: list[dict], caminho: str):
    """Salva o relatório em Excel com layout profissional (3 abas + gráfico)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from collections import Counter
    from datetime import datetime, date

    AZUL_MARINHO = '1F3864'
    AZUL_SECUND  = '2E5496'
    LINHA_PAR    = 'EBF3FB'
    HOJE         = date.today()
    AGORA        = datetime.now().strftime('%d/%m/%Y às %H:%M')

    SITUACAO_MAP = {
        'integrado':               'CONCLUIDO',
        'pago':                    'PAGO',
        'emissão de cartão':       'EM EMISSAO DE CARTAO',
        'averbação':               'AGUARD. AVERBACAO',
        'integração documental':   'DOC. EM ANDAMENTO',
        'formalização digital':    'FORMALIZANDO',
        'formalização interna':    'FORMALIZANDO',
        'aprovação corban':        'AGUARD. APROVACAO',
        'in100':                   'ANALISE IN100',
        'crédito':                 'ANALISE CREDITO',
        'protocolar documentação': 'PROTOCOLO PENDENTE',
        'lançamento do título':    'AGUARD. LANCAMENTO',
        'cancelado':               'CANCELADO',
        'não encontrado':          'VERIFICAR MANUAL.',
        'não identificado':        'VERIFICAR MANUAL.',
    }
    SITUACAO_CORES = {
        'CONCLUIDO':            ('C6EFCE', '276221'),
        'PAGO':                 ('FFEB9C', '9C5700'),
        'EM EMISSAO DE CARTAO': ('DDEEFF', '003399'),
        'AGUARD. AVERBACAO':    ('FFCC99', 'C55A11'),
        'DOC. EM ANDAMENTO':    ('FFD7D7', '9C0006'),
        'FORMALIZANDO':         ('E8D5F5', '5B2D8E'),
        'AGUARD. APROVACAO':    ('D5F5EE', '1A7B5E'),
        'ANALISE IN100':        ('D5F0FA', '0070C0'),
        'ANALISE CREDITO':      ('FCE4D6', 'C55A11'),
        'PROTOCOLO PENDENTE':   ('DDD9EE', '4B3E99'),
        'AGUARD. LANCAMENTO':   ('F5DFD0', '843C0C'),
        'CANCELADO':            ('D9D9D9', '595959'),
        'VERIFICAR MANUAL.':    ('FF0000', 'FFFFFF'),
    }
    FASES_CONCLUIDAS = {'integrado', 'pago'}
    FASES_PENDENCIA  = {'cancelado', 'não encontrado', 'não identificado'}

    def borda():
        s = Side(style='thin', color='B0C4DE')
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(cor):
        return PatternFill('solid', fgColor=cor)

    def dias_desde(data_str: str):
        try:
            d = datetime.strptime(data_str.strip(), '%d/%m/%Y').date()
            return (HOJE - d).days
        except Exception:
            return None

    wb = Workbook()

    # ── ABA 1: RELATÓRIO DETALHADO ───────────────────────────────────────
    ws = wb.active
    ws.title = 'Relatorio'
    cabecalhos = ['Seq.', 'Numero do Contrato', 'Fase da Esteira',
                  'Situacao Resumida', 'Data de Averbacao', 'Dias desde Averbacao']
    larguras   = [8, 24, 26, 22, 20, 22]

    for ci, texto in enumerate(cabecalhos, start=1):
        c = ws.cell(row=1, column=ci, value=texto)
        c.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        c.fill      = fill(AZUL_MARINHO)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = borda()
    ws.row_dimensions[1].height = 26

    for ri, item in enumerate(dados, start=2):
        fase   = item['Fase da Esteira']
        averb  = item.get('Averbação', item.get('Averbacao', ''))
        chave  = fase.lower().strip()
        sit    = SITUACAO_MAP.get(chave, 'EM PROCESSO')
        dias   = dias_desde(averb)
        cor_linha = LINHA_PAR if ri % 2 == 0 else 'FFFFFF'
        num    = item.get('Número_Contrato', item.get('Numero_Contrato', ''))

        valores = [item['Sequencial'], num, fase, sit, averb, dias if dias is not None else '-']

        for ci, val in enumerate(valores, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = borda()
            if ci == 3:
                bg, fg = CORES_FASE.get(chave, (cor_linha, '000000'))
                c.fill = fill(bg); c.font = Font(name='Calibri', size=10, bold=True, color=fg)
            elif ci == 4:
                sb, sf = SITUACAO_CORES.get(sit, (cor_linha, '000000'))
                c.fill = fill(sb); c.font = Font(name='Calibri', size=10, bold=True, color=sf)
            elif ci == 6 and dias is not None:
                if dias <= 7:
                    db, df = 'C6EFCE', '276221'
                elif dias <= 30:
                    db, df = 'FFEB9C', '9C5700'
                else:
                    db, df = 'FFD7D7', '9C0006'
                c.fill = fill(db); c.font = Font(name='Calibri', size=10, bold=True, color=df)
            else:
                c.fill = fill(cor_linha); c.font = Font(name='Calibri', size=10)
        ws.row_dimensions[ri].height = 18

    for ci, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = larg
    ws.freeze_panes = 'A2'

    # ── ABA 2: RESUMO ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Resumo')
    ws2.merge_cells('A1:D1')
    t = ws2['A1']
    t.value     = 'RESUMO — CONTRATOS POR FASE DA ESTEIRA'
    t.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
    t.fill      = fill(AZUL_MARINHO)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells('A2:D2')
    g = ws2['A2']
    g.value     = f'Gerado em: {AGORA}'
    g.font      = Font(name='Calibri', italic=True, color='595959', size=10)
    g.fill      = fill('F2F7FC')
    g.alignment = Alignment(horizontal='right', vertical='center')
    ws2.row_dimensions[2].height = 18

    for ci, texto in enumerate(['Fase da Esteira', 'Qtd. Contratos', '% do Total'], start=1):
        c = ws2.cell(row=3, column=ci, value=texto)
        c.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        c.fill      = fill(AZUL_SECUND)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = borda()
    ws2.row_dimensions[3].height = 22

    contagem  = Counter(d['Fase da Esteira'] for d in dados)
    total     = len(dados)
    fases_ord = sorted(contagem.items(), key=lambda x: x[1], reverse=True)

    for ri, (fase, qtd) in enumerate(fases_ord, start=4):
        pct   = f"{qtd / total * 100:.1f}%"
        chave = fase.lower().strip()
        bg, fg = CORES_FASE.get(chave, ('F2F7FC', '000000'))

        c1 = ws2.cell(row=ri, column=1, value=fase)
        c1.fill = fill(bg); c1.font = Font(name='Calibri', size=11, bold=True, color=fg)
        c1.alignment = Alignment(horizontal='left', vertical='center', indent=1); c1.border = borda()

        c2 = ws2.cell(row=ri, column=2, value=qtd)
        c2.fill = fill('FFFFFF'); c2.font = Font(name='Calibri', size=11, bold=True, color=AZUL_MARINHO)
        c2.alignment = Alignment(horizontal='center', vertical='center'); c2.border = borda()

        c3 = ws2.cell(row=ri, column=3, value=pct)
        c3.fill = fill('FFFFFF'); c3.font = Font(name='Calibri', size=11, color='595959')
        c3.alignment = Alignment(horizontal='center', vertical='center'); c3.border = borda()
        ws2.row_dimensions[ri].height = 22

    rt = len(contagem) + 4
    for ci, val in enumerate(['TOTAL GERAL', total, '100%'], start=1):
        c = ws2.cell(row=rt, column=ci, value=val)
        c.fill = fill(AZUL_MARINHO)
        c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='left' if ci == 1 else 'center', vertical='center', indent=1 if ci == 1 else 0)
        c.border = borda()
    ws2.row_dimensions[rt].height = 22
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 14

    try:
        chart = BarChart()
        chart.type = 'col'
        chart.title = 'Contratos por Fase da Esteira'
        chart.y_axis.title = 'Quantidade'
        chart.x_axis.title = 'Fase'
        chart.style  = 10
        chart.width  = 20
        chart.height = 14
        data_ref = Reference(ws2, min_col=2, min_row=3, max_row=3 + len(contagem))
        cats_ref = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(contagem))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws2.add_chart(chart, 'E3')
    except Exception:
        pass

    # ── ABA 3: PENDÊNCIAS ────────────────────────────────────────────────
    ws3 = wb.create_sheet('Pendencias')
    ws3.merge_cells('A1:F1')
    p = ws3['A1']
    p.value     = 'PENDENCIAS — CONTRATOS QUE REQUEREM ATENCAO'
    p.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
    p.fill      = fill('C00000')
    p.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 30

    cab3 = ['Seq.', 'Numero do Contrato', 'Fase da Esteira', 'Situacao', 'Data Averbacao', 'Motivo']
    for ci, texto in enumerate(cab3, start=1):
        c = ws3.cell(row=3, column=ci, value=texto)
        c.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        c.fill      = fill('C00000')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = borda()
    ws3.row_dimensions[3].height = 22

    ri3 = 4
    for item in dados:
        fase  = item['Fase da Esteira']
        chave = fase.lower().strip()
        averb = item.get('Averbação', item.get('Averbacao', ''))
        num   = item.get('Número_Contrato', item.get('Numero_Contrato', ''))

        if chave in FASES_PENDENCIA:
            motivo = 'Nao localizado no sistema' if 'encontrado' in chave or 'identificado' in chave else 'Contrato cancelado'
        elif chave not in FASES_CONCLUIDAS:
            motivo = f'Em andamento: {fase}'
        else:
            continue

        sit = SITUACAO_MAP.get(chave, 'EM PROCESSO')
        cor_linha = 'FFF2F2' if ri3 % 2 == 0 else 'FFFFFF'
        vals3 = [item['Sequencial'], num, fase, sit, averb, motivo]
        for ci, val in enumerate(vals3, start=1):
            c = ws3.cell(row=ri3, column=ci, value=val)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = borda()
            if ci == 3:
                bg, fg = CORES_FASE.get(chave, (cor_linha, '000000'))
                c.fill = fill(bg); c.font = Font(name='Calibri', size=10, bold=True, color=fg)
            elif ci == 4:
                sb, sf = SITUACAO_CORES.get(sit, (cor_linha, '000000'))
                c.fill = fill(sb); c.font = Font(name='Calibri', size=10, bold=True, color=sf)
            else:
                c.fill = fill(cor_linha); c.font = Font(name='Calibri', size=10)
        ws3.row_dimensions[ri3].height = 18
        ri3 += 1

    total_pend = ri3 - 4
    ws3.merge_cells('A2:F2')
    g3 = ws3['A2']
    g3.value     = f'Gerado em: {AGORA}  |  Total de pendencias: {total_pend}'
    g3.font      = Font(name='Calibri', italic=True, color='595959', size=10)
    g3.fill      = fill('FFF2F2')
    g3.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws3.row_dimensions[2].height = 18

    for ci, larg in enumerate([8, 24, 26, 22, 20, 30], start=1):
        ws3.column_dimensions[get_column_letter(ci)].width = larg
    ws3.freeze_panes = 'A4'

    wb.save(caminho)
    print(f"\n✓ Relatório salvo em: {caminho}")


# ============================================================
# MAIN
# ============================================================
def main():
    print("="*60)
    print(" AUTOMAÇÃO — VALIDAÇÃO REGULATÓRIA DATAPREV")
    print("="*60)

    contratos = extrair_contratos(ARQUIVO_GZ)
    driver    = iniciar_navegador()
    resultados = []

    print("\n" + "="*60)
    print(f"ETAPA 3 — Consultando {len(contratos)} contratos")
    print("="*60)

    try:
        for seq, contrato in enumerate(contratos, start=1):
            print(f"\n[{seq:03d}/{len(contratos)}] Contrato: {contrato}")
            fase, averbacao = buscar_proposta(driver, contrato)

            resultados.append({
                'Sequencial'      : seq,
                'Número_Contrato' : contrato,
                'Fase da Esteira' : fase,
                'Averbação'       : averbacao,
            })

            if seq % 10 == 0:
                salvar_excel(resultados, ARQUIVO_SAIDA.replace('.xlsx', '_parcial.xlsx'))
                print(f"  ✓ Salvo parcial ({seq} contratos)")

            time.sleep(PAUSA_ENTRE_CONTRATOS)

    finally:
        driver.quit()
        print("\n✓ Navegador fechado")

    if resultados:
        salvar_excel(resultados, ARQUIVO_SAIDA)
        df_final = pd.DataFrame(resultados)
        print("\n" + "="*60)
        print("RESUMO FINAL")
        print("="*60)
        print(f"Total processados : {len(df_final)}")
        print(f"Não encontrados   : {(df_final['Fase da Esteira'] == 'NÃO ENCONTRADO').sum()}")
        print(f"Com averbação     : {(df_final['Averbação'] != '').sum()}")
    else:
        print("\n✗ Nenhum resultado obtido.")

    try:
        input("\nPressione ENTER para sair...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()
